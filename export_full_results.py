"""Export complete trial data from MASID experiments.

Exports:
1. trials_summary.csv — scores and metadata per trial
2. agent_outputs_full.csv — every agent response from every round
3. trial_details/ — one text file per trial with full readable output
4. summary_stats.txt — aggregate statistics
5. masid_full_export.xlsx — Excel workbook with all data in multiple sheets
"""

import sqlite3
import json
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_PATH = "data/experiments.db"
EXPORT_DIR = Path("data/export")

HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")


def main():
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    (EXPORT_DIR / "trial_details").mkdir(exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    trials = conn.execute("SELECT * FROM trials ORDER BY architecture, domain, task_id").fetchall()
    print(f"Total trials: {len(trials)}")

    # 1. Trial summary CSV
    with open(EXPORT_DIR / "trials_summary.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "trial_id", "architecture", "domain", "model", "task_id", "seed",
            "quality_score", "correctness", "completeness", "coherence", "integration",
            "total_tokens", "total_latency_seconds", "num_rounds",
            "duplication_rate", "fault_type", "fault_agent",
            "judge_rationale", "metadata_json",
        ])
        for t in trials:
            writer.writerow([
                t["trial_id"], t["architecture"], t["domain"], t["model"],
                t["task_id"], t["seed"],
                t["quality_score"], t["correctness"], t["completeness"],
                t["coherence"], t["integration"],
                t["total_tokens"], t["total_latency_seconds"], t["num_rounds"],
                t["duplication_rate"], t["fault_type"], t["fault_agent"],
                t["judge_rationale"], t["metadata_json"],
            ])
    print(f"  Exported: {EXPORT_DIR / 'trials_summary.csv'}")

    # 2. Full agent outputs CSV
    outputs = conn.execute("""
        SELECT ao.trial_id, t.architecture, t.domain, t.task_id, t.seed,
               ao.agent_id, ao.role, ao.round_number,
               ao.content, ao.prompt_tokens, ao.completion_tokens, ao.latency_seconds
        FROM agent_outputs ao
        JOIN trials t ON ao.trial_id = t.trial_id
        ORDER BY t.architecture, t.domain, t.task_id, ao.trial_id, ao.round_number, ao.role
    """).fetchall()

    with open(EXPORT_DIR / "agent_outputs_full.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([
            "trial_id", "architecture", "domain", "task_id", "seed",
            "agent_id", "role", "round_number",
            "content", "prompt_tokens", "completion_tokens", "latency_seconds",
        ])
        for o in outputs:
            writer.writerow([
                o["trial_id"], o["architecture"], o["domain"], o["task_id"], o["seed"],
                o["agent_id"], o["role"], o["round_number"],
                o["content"], o["prompt_tokens"], o["completion_tokens"], o["latency_seconds"],
            ])
    print(f"  Exported: {EXPORT_DIR / 'agent_outputs_full.csv'}")

    # 3. Per-trial detail text files
    for t in trials:
        tid = t["trial_id"]
        filename = f"{t['architecture']}_{t['domain']}_{t['task_id']}_{tid}.txt"
        filepath = EXPORT_DIR / "trial_details" / filename

        trial_outputs = conn.execute("""
            SELECT role, round_number, content, prompt_tokens, completion_tokens, latency_seconds
            FROM agent_outputs WHERE trial_id = ?
            ORDER BY round_number, role
        """, (tid,)).fetchall()

        with open(filepath, "w") as f:
            f.write(f"{'=' * 80}\n")
            f.write(f"TRIAL: {tid}\n{'=' * 80}\n")
            f.write(f"Architecture:  {t['architecture']}\nDomain:        {t['domain']}\n")
            f.write(f"Task:          {t['task_id']}\nModel:         {t['model']}\nSeed:          {t['seed']}\n\n")
            f.write(f"SCORES:\n  Quality: {t['quality_score']}\n  Correctness: {t['correctness']}\n")
            f.write(f"  Completeness: {t['completeness']}\n  Coherence: {t['coherence']}\n  Integration: {t['integration']}\n\n")
            f.write(f"EFFICIENCY:\n  Tokens: {t['total_tokens']}\n  Latency: {t['total_latency_seconds']}s\n  Rounds: {t['num_rounds']}\n\n")
            f.write(f"JUDGE RATIONALE:\n  {t['judge_rationale'] or '(empty)'}\n\n")

            if t["metadata_json"]:
                try:
                    meta = json.loads(t["metadata_json"])
                    if "execution_score" in meta:
                        f.write(f"CODE EXECUTION:\n  Exec score: {meta.get('execution_score')}\n")
                        f.write(f"  Syntax: {meta.get('syntax_valid')}  Runs: {meta.get('code_runs')}\n")
                        f.write(f"  Tests: {meta.get('tests_passed')}/{meta.get('tests_total')}\n\n")
                except (json.JSONDecodeError, TypeError):
                    pass

            rounds = defaultdict(list)
            for o in trial_outputs:
                rounds[o["round_number"]].append(o)

            for round_num in sorted(rounds.keys()):
                f.write(f"\n{'#' * 80}\n# ROUND {round_num + 1}\n{'#' * 80}\n")
                for o in rounds[round_num]:
                    f.write(f"\n{'-' * 60}\n{o['role']} (round {round_num + 1})\n")
                    f.write(f"Tokens: {o['prompt_tokens']} in / {o['completion_tokens']} out | {o['latency_seconds']}s\n")
                    f.write(f"{'-' * 60}\n{o['content']}\n")

    print(f"  Exported: {len(trials)} trial detail files")

    # 4. Summary statistics
    groups = defaultdict(list)
    for t in trials:
        groups[(t["architecture"], t["domain"])].append(t)

    task_groups = defaultdict(list)
    for t in trials:
        task_groups[(t["domain"], t["task_id"])].append(t)

    with open(EXPORT_DIR / "summary_stats.txt", "w") as f:
        f.write(f"MASID EXPERIMENT RESULTS — SUMMARY STATISTICS\nTotal trials: {len(trials)}\n{'=' * 90}\n\n")
        f.write(f"{'Arch':<6} {'Domain':<22} {'N':>3} {'Avg Q':>7} {'Min':>6} {'Max':>6} {'StdDev':>7} {'Avg Tok':>8} {'Avg Lat':>8}\n")
        f.write("-" * 90 + "\n")
        for (arch, domain) in sorted(groups.keys(), key=lambda x: (x[1], x[0])):
            g = groups[(arch, domain)]
            scores = [t["quality_score"] for t in g]
            n = len(scores)
            avg = sum(scores) / n
            std = (sum((s - avg) ** 2 for s in scores) / n) ** 0.5
            f.write(f"{arch:<6} {domain:<22} {n:>3} {avg:>7.3f} {min(scores):>6.3f} {max(scores):>6.3f} {std:>7.3f} {sum(t['total_tokens'] for t in g)/n:>8.0f} {sum(t['total_latency_seconds'] for t in g)/n:>8.1f}\n")

        f.write(f"\n{'=' * 90}\nPER-TASK COMPARISON\n{'=' * 90}\n\n")
        for (domain, task_id) in sorted(task_groups.keys()):
            f.write(f"\n  {domain} / {task_id}:\n")
            arch_scores = defaultdict(list)
            for t in task_groups[(domain, task_id)]:
                arch_scores[t["architecture"]].append(t["quality_score"])
            for arch in ["irm", "jro", "iamd"]:
                scores = arch_scores.get(arch, [])
                if scores:
                    f.write(f"    {arch:<6}: {' '.join(f'{s:.2f}' for s in scores)}  avg={sum(scores)/len(scores):.3f}\n")

        f.write(f"\n{'=' * 90}\nDOMAIN WINNERS\n{'=' * 90}\n\n")
        for domain in ["software_dev", "research_synthesis", "project_planning"]:
            avgs = {}
            tokens = {}
            for arch in ["irm", "jro", "iamd"]:
                g = groups.get((arch, domain), [])
                if g:
                    avgs[arch] = sum(t["quality_score"] for t in g) / len(g)
                    tokens[arch] = sum(t["total_tokens"] for t in g) / len(g)
            if avgs:
                winner = max(avgs, key=avgs.get)
                f.write(f"  {domain}:\n")
                for arch in sorted(avgs, key=avgs.get, reverse=True):
                    marker = " <- BEST" if arch == winner else ""
                    f.write(f"    {arch:<6} avg={avgs[arch]:.3f}  tokens={tokens[arch]:,.0f}{marker}\n")
                f.write("\n")

    print(f"  Exported: {EXPORT_DIR / 'summary_stats.txt'}")

    # 5. Combined Excel workbook
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    headers = ["Architecture", "Domain", "N", "Avg Quality", "Min", "Max", "StdDev", "Avg Tokens", "Avg Latency"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
    row = 2
    for (arch, domain) in sorted(groups.keys(), key=lambda x: (x[1], x[0])):
        g = groups[(arch, domain)]
        scores = [t["quality_score"] for t in g]
        n = len(scores)
        avg = sum(scores) / n
        std = (sum((s - avg) ** 2 for s in scores) / n) ** 0.5
        for col, val in enumerate([arch, domain, n, round(avg, 3), round(min(scores), 3),
                                    round(max(scores), 3), round(std, 3),
                                    round(sum(t["total_tokens"] for t in g) / n),
                                    round(sum(t["total_latency_seconds"] for t in g) / n, 1)], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Sheet 2: All Trials
    ws2 = wb.create_sheet("All Trials")
    t_headers = ["Trial ID", "Architecture", "Domain", "Task", "Seed", "Quality",
                 "Correctness", "Completeness", "Coherence", "Integration",
                 "Tokens", "Latency (s)", "Rounds", "Judge Rationale"]
    for col, h in enumerate(t_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
    for i, t in enumerate(trials, 2):
        vals = [t["trial_id"], t["architecture"], t["domain"], t["task_id"], t["seed"],
                t["quality_score"], t["correctness"], t["completeness"],
                t["coherence"], t["integration"], t["total_tokens"],
                t["total_latency_seconds"], t["num_rounds"], t["judge_rationale"] or ""]
        for col, val in enumerate(vals, 1):
            ws2.cell(row=i, column=col, value=val)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["N"].width = 50
    ws2.auto_filter.ref = f"A1:N{len(trials) + 1}"

    # Sheet 3: Trial Details (final round agent outputs)
    ws3 = wb.create_sheet("Trial Details")
    d_headers = ["Trial ID", "Architecture", "Domain", "Task", "Quality",
                 "Role", "Round", "Output", "Prompt Tokens", "Completion Tokens"]
    for col, h in enumerate(d_headers, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL

    final_outputs = conn.execute("""
        SELECT ao.trial_id, t.architecture, t.domain, t.task_id, t.quality_score,
               ao.role, ao.round_number, ao.content, ao.prompt_tokens, ao.completion_tokens
        FROM agent_outputs ao
        JOIN trials t ON ao.trial_id = t.trial_id
        WHERE ao.round_number = (t.num_rounds - 1)
        ORDER BY t.architecture, t.domain, t.task_id, ao.trial_id, ao.role
    """).fetchall()

    for i, o in enumerate(final_outputs, 2):
        content = (o["content"] or "")[:32000]
        vals = [o["trial_id"], o["architecture"], o["domain"], o["task_id"],
                o["quality_score"], o["role"], o["round_number"] + 1, content,
                o["prompt_tokens"], o["completion_tokens"]]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.alignment = WRAP_ALIGN if col == 8 else TOP_ALIGN
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["F"].width = 18
    ws3.column_dimensions["H"].width = 80
    ws3.auto_filter.ref = f"A1:J{len(final_outputs) + 1}"

    # Sheet 4: Per-Task Comparison
    ws4 = wb.create_sheet("Per-Task")
    pt_headers = ["Domain", "Task", "Architecture", "Avg Quality", "N", "Scores"]
    for col, h in enumerate(pt_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
    row = 2
    for (domain, task_id) in sorted(task_groups.keys()):
        arch_scores = defaultdict(list)
        for t in task_groups[(domain, task_id)]:
            arch_scores[t["architecture"]].append(t["quality_score"])
        for arch in ["irm", "jro", "iamd"]:
            scores = arch_scores.get(arch, [])
            if scores:
                ws4.cell(row=row, column=1, value=domain)
                ws4.cell(row=row, column=2, value=task_id)
                ws4.cell(row=row, column=3, value=arch)
                ws4.cell(row=row, column=4, value=round(sum(scores) / len(scores), 3))
                ws4.cell(row=row, column=5, value=len(scores))
                ws4.cell(row=row, column=6, value=", ".join(f"{s:.2f}" for s in scores))
                row += 1
    for col in range(1, 7):
        ws4.column_dimensions[get_column_letter(col)].width = 20
    ws4.column_dimensions["F"].width = 60

    xlsx_path = EXPORT_DIR / "masid_full_export.xlsx"
    wb.save(xlsx_path)
    print(f"  Exported: {xlsx_path}")

    conn.close()
    print(f"\nAll exports complete in {EXPORT_DIR}/")


if __name__ == "__main__":
    main()